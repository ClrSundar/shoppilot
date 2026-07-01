#!/usr/bin/env python3
"""
Generate a sample bulk upload Excel template.
Run: python3 generate_bulk_upload_template.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

def create_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Bulk Upload"

    headers = [
        "type",
        "name",
        "description",
        "category",
        "costPrice",
        "sellingPrice",
        "sku",
        "brand",
        "unit",
        "imageUrl",
        "openingStock",
        "reorderLevel",
    ]
    ws.append(headers)

    sample_rows = [
        ["category", "Electronics", "Electronic devices and gadgets", "", "", "", "", "", "", "", "", ""],
        ["category", "Hardware", "Tools and equipment", "", "", "", "", "", "", "", "", ""],
        ["product", "Laptop", "", "Electronics", 500, 750, "LAPTOP-001", "Dell", "NOS", "", "", ""],
        ["product", "Monitor", "", "Electronics", 150, 200, "MON-001", "HP", "NOS", "", "", ""],
        ["product", "Hammer", "", "Hardware", 10, 15, "HAMMER-001", "Stanley", "NOS", "", "", ""],
        ["inventory", "Laptop", "", "", "", "", "LAPTOP-001", "", "", "", 10, 5],
        ["inventory", "Monitor", "", "", "", "", "MON-001", "", "", "", 20, 8],
        ["inventory", "Hammer", "", "", "", "", "HAMMER-001", "", "", "", 50, 20],
    ]
    for row in sample_rows:
        ws.append(row)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2
    
    # Save the workbook
    template_path = Path(__file__).parent / "bulk_upload_template.xlsx"
    wb.save(template_path)
    print(f"✅ Template created: {template_path}")
    print(f"\nThe template includes:")
    print("  • 1 sheet named 'Bulk Upload'")
    print("  • Use the 'type' column to mark rows as category, product, or inventory")
    print("\nReplace sample data with your actual data and upload to POST /bulk-upload")

if __name__ == "__main__":
    create_template()
